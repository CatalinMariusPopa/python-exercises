from tkinter import *
from win32com import client
from openpyxl.chart import (PieChart, Reference)
import openpyxl

root = Tk()
root.geometry("800x400")
root.resizable(0,0)
root.title("Test Cases Parser")
Label(root, text="Test Cases Parser", font="arial 15 bold").pack()


tester_label = Label(root, text="Tester Name", font="arial 10 bold").pack()
tester_str = StringVar()
Entry(root, textvariable=tester_str).pack()

values = [0,0,0]
#totalTestCases_counter = 0
#failTestCases_counter = 0
#passTestCases_counter = 0

path = r'C:\Users\Catalin\Desktop\examen\Test_Case_Format_(Popa Marius Catalin) (Project Arduido PingPong).xlsx'
path_PDF = r'C:\Users\Catalin\Desktop\examen\Test_Case_Format_(Popa Marius Catalin) (Project Arduido PingPong).pdf'

def compareValues():
   # totalTestCases_counter = 0
   # failTestCases_counter = 0
   # passTestCases_counter = 0
    wb = openpyxl.load_workbook(path, read_only=False)
    first_sheet = wb.get_sheet_by_name('test case format')

    for i in range(1, first_sheet.max_row):
        if (first_sheet.cell(row=i, column=7).value) == 'fail':
            values[0] = values[0]+1
        elif (first_sheet.cell(row=i, column=7).value) == 'pass':
            values[1] = values[1] + 1


    values[2] = values[0] + values[1]
    #print('Total teste fail: ',  values[0])
    #print('Total teste pass: ',  values[1])
    #print('Total teste: ', values[2])




def generateChart():
    wb = openpyxl.load_workbook(path, read_only=False)
    testCasesSheet = wb.get_sheet_by_name('test case format')
    testerName = testCasesSheet['E1'].value
    try:
        reportSheet = wb.get_sheet_by_name("Report")
    except:
        wb.create_sheet("Report")
        reportSheet = wb.get_sheet_by_name("Report")
    reportSheet['A1'] = 'Tester ID:'
    reportSheet['B1'] = testerName
    reportSheet['A2'] = 'Failed Test Cases:'
    reportSheet['B2'] = values[0]
    reportSheet['A3'] = 'Pass Test Cases:'
    reportSheet['B3'] = values[1]
    reportSheet['A4'] = 'Total Number of Test Cases:'
    reportSheet['B4'] = values[2]
    wb.save(path)
    createChart()
    # Open Microsoft Excel
    excel = client.Dispatch("Excel.Application")

    # Read Excel File
    sheets = excel.Workbooks.Open(path)
    work_sheets = sheets.Worksheets[2]

    # Convert into PDF File
    work_sheets.ExportAsFixedFormat(0, path_PDF)

def createChart():
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb.get_sheet_by_name('Report')
    pie = PieChart()

    labels = Reference(sheet, min_col=1, min_row=2, max_row=3)
    data = Reference(sheet, min_col=2, min_row=2, max_row=3)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = 'Test Cases'

    pie.width = 14
    pie.lenght = 7

    sheet.add_chart(pie, 'A6')

    wb.save(path)


def buttonPressed():
    compareValues()
    generateChart()
Button(root, text="Generate report", command=buttonPressed).pack(pady=10)


root.mainloop()




